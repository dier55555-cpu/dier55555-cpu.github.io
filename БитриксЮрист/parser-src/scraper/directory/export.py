"""Выгрузка справочника в JSON и Excel (листы по типу суда + города/районы)."""

from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .normalize import CourtRecord

COLUMNS = [
    ("code", "Код"),
    ("name", "Суд"),
    ("court_type", "Тип"),
    ("court_type_name", "Тип (название)"),
    ("region_code", "Код региона"),
    ("region", "Регион"),
    ("city", "Город / населённый пункт"),
    ("district", "Район (из названия суда)"),
    ("address", "Адрес"),
    ("website", "Сайт (как в DaData)"),
    ("sudrf_domain", "Домен sudrf (живой)"),
    ("parser_supported", "Парсер G1/U1"),
]

TYPE_SHEETS = {
    "RS": "Районные_городские",
    "OS": "Областные",
    "MS": "Мировые",
    "AJ": "Апелляция",
    "KJ": "Кассация",
    "VS": "Верховный_суд",
    "AS": "Арбитраж_субъектов",
    "AA": "Арбитраж_апелляция",
    "AO": "Арбитраж_округа",
    "AI": "СИП",
}


def _sheet_title(name: str) -> str:
    cleaned = "".join("_" if ch in r":\/?*[]" else ch for ch in name).strip()
    return (cleaned or "Лист")[:31]


def _write_header(ws: Worksheet, headers: Sequence[str]) -> None:
    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"


def _autosize(ws: Worksheet, max_width: int = 48) -> None:
    for index, column in enumerate(ws.columns, start=1):
        width = 10
        for cell in column[:80]:
            width = max(width, min(max_width, len(str(cell.value or "")) + 2))
        ws.column_dimensions[get_column_letter(index)].width = width


def _row_values(record: CourtRecord) -> list[object]:
    data = record.as_dict()
    values = []
    for key, _title in COLUMNS:
        value = data.get(key)
        if key == "parser_supported":
            values.append("да" if value else "нет")
        else:
            values.append(value)
    return values


def write_json(path: Path, records: Sequence[CourtRecord], *, extra: dict | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    counts: dict[str, int] = defaultdict(int)
    for record in records:
        counts[record.court_type or "?"] += 1
    payload = {
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "count": len(records),
        "counts": dict(sorted(counts.items())),
        "courts": [record.as_dict() for record in records],
    }
    if extra:
        payload.update(extra)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_xlsx(path: Path, records: Sequence[CourtRecord]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    all_sheet = wb.active
    all_sheet.title = _sheet_title("Все_суды")
    _fill_courts_sheet(all_sheet, records)

    by_type: dict[str, list[CourtRecord]] = defaultdict(list)
    for record in records:
        by_type[record.court_type or "?"].append(record)
    for court_type, title in TYPE_SHEETS.items():
        rows = by_type.get(court_type) or []
        if not rows:
            continue
        ws = wb.create_sheet(_sheet_title(title))
        _fill_courts_sheet(ws, rows)

    cities = wb.create_sheet(_sheet_title("Города"))
    _fill_cities_sheet(cities, records)

    districts = wb.create_sheet(_sheet_title("Районы"))
    _fill_districts_sheet(districts, records)

    wb.save(path)


def _fill_courts_sheet(ws: Worksheet, records: Sequence[CourtRecord]) -> None:
    _write_header(ws, [title for _key, title in COLUMNS])
    for record in records:
        ws.append(_row_values(record))
    _autosize(ws)


def _fill_cities_sheet(ws: Worksheet, records: Sequence[CourtRecord]) -> None:
    _write_header(ws, ["Регион", "Город / населённый пункт", "Судов", "Названия судов"])
    grouped: dict[tuple[str, str], list[CourtRecord]] = defaultdict(list)
    for record in records:
        grouped[(record.region or "—", record.city or "—")].append(record)
    for (region, city) in sorted(grouped):
        rows = grouped[(region, city)]
        names = "; ".join(item.name for item in rows)
        ws.append([region, city, len(rows), names])
    _autosize(ws, max_width=60)


def _fill_districts_sheet(ws: Worksheet, records: Sequence[CourtRecord]) -> None:
    _write_header(ws, ["Регион", "Город", "Район", "Суд", "Домен sudrf", "Парсер"])
    district_rows = [r for r in records if r.district]
    district_rows.sort(key=lambda r: (r.region, r.city, r.district, r.name))
    for record in district_rows:
        ws.append([
            record.region,
            record.city,
            record.district,
            record.name,
            record.sudrf_domain,
            "да" if record.parser_supported else "нет",
        ])
    _autosize(ws)


def sorted_records(records: Iterable[CourtRecord]) -> list[CourtRecord]:
    return sorted(
        records,
        key=lambda r: (r.region, r.city, r.district, r.court_type, r.name, r.code),
    )
